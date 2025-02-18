import pyshark
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

frames = {}

# for packet in pyshark.FileCapture("250203.pcapng"):
#   sIndex                    = packet.udp.stream
  
#   if not sIndex in frames:
#     frames[sIndex] = {
#       "eth"                : {},
#       "ip"                 : {},
#       "packets"            : [],
#       "stream_index"       : sIndex,
#       }
    
#   _frame                    = frames[sIndex]
#   _frame["packets"].append(packet)
  
#   if not packet.eth.src in _frame["eth"]:
#     _frame["eth"][packet.eth.src] = True
#   if not packet.eth.dst in _frame["eth"]:
#     _frame["eth"][packet.eth.dst] = True
    
#   if not packet.ip.src in _frame["ip"]:
#     _frame["ip"][packet.ip.src] = True
#   if not packet.ip.dst in _frame["ip"]:
#     _frame["ip"][packet.ip.dst] = True
 
# print(frames["100"])

wb = Workbook()
ws = wb.active

ws["A1"] = "Devices_Name"
ws["B1"] = "Origin_IP"
ws["C1"] = "Routed_IP"
ws["D1"] = "Destination_Device_IP"
ws["E1"] = "Network_Number"
ws["F1"] = "WhoIs_Count"
ws["G1"] = "WhoHas_Count"
ws["H1"] = "IAm_Count"
ws["I1"] = "Who_Is_Router_Count"
ws["J1"] = "Who_Is_Router_Global_Count"
ws["K1"] = "I_Am_Router_Count"

column = "A"
row = "2"
ws[column + row] = "hello"
table = Table(displayName="Table", ref="A1:A2")
style = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=False
)
table.tableStyleInfo = style
ws.add_table(table)
wb.save("pcap_sheet.xlsx")